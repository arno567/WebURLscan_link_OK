# coding=utf-8
#python3.7
import time
import requests
from openpyxl import load_workbook
import threading
from concurrent.futures import ThreadPoolExecutor
import random

#解决https报错
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

right_code = ['200','304']

data_msg = time.strftime('%m{m}%d{d}%H{h}%M{f}{s}').format(m='月', d='日', h='时', f='分', s='异常访问情况')

def getheaders():#随机获取头信息
    user_agent_list = [ \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1" \
        "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6", \
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1", \
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5", \
        "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3", \
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.0 Safari/536.3", \
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24", \
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24"
    ]
    UserAgent=random.choice(user_agent_list)
    headers = {'User-Agent': UserAgent}
    return headers

def get_real_url(url,try_count = 1):#重定向获取真实url
    if try_count > 3:
        return url
    try:
        headers = getheaders()
        rs = requests.get(url,headers=headers,verify=False,timeout=5)
        if rs.status_code > 400:
            return get_real_url(url,try_count+1)
        return rs.url
    except:
        return get_real_url(url, try_count + 1)

def toxlsx(all,dirname):#结果写入xlsx
    all_cells = all
    wb = load_workbook(dirname)
    sheets = wb.worksheets  # 获取当前所有的sheet
    for sheet in sheets:
        for cell in all_cells:
            title = sheet.title
            if cell:
                if cell[0] in title:
                    rowNum = int(cell[1])
                    NewcolumnMaxNum = int(cell[2]) + 1
                    sheet.cell(row=1,column=NewcolumnMaxNum, value=data_msg)
                    if cell[4] in "异常":
                        sheet.cell(row=rowNum, column=NewcolumnMaxNum, value=cell[4])
                else:
                    continue
            else:
                continue
    wb.save(filename=dirname)
    wb.close()

def doxlsx(dirname):#获取url列
    wb = load_workbook(dirname)
    sheets = wb.worksheets  # 获取当前所有的sheet
    #print(sheets)
    all_url = []

    print("获取所有URL...")
    for sheet in sheets:

        #columns = sheet.columns
        sheet_title =  sheet.title #sheet名称
        rowMaxNum = sheet.max_row #行数
        columnsMaxNum = sheet.max_column #列数
        #print(sheet.cell(row=2,column=3).value) #获取单元格数据
        url_colum_num = '' #url存放的第几列

        num = 0
        for i in range(1,columnsMaxNum+1):#获取url在第几列
            url_colums = str(sheet.cell(row=1, column=i).value)
            if url_colums:
                if url_colums in "网站域名|域名|URL|url|域名/URL|域名/url|链接":
                    url_colum_num = int(i)
                num += 1

        if url_colum_num:
            for i in range(1,rowMaxNum+1):
                url = str(sheet.cell(row=i, column=url_colum_num).value)
                if url is not 'none':
                    if '.' in url:
                        a = []
                        #print(url)
                        a.append(sheet_title)
                        a.append(str(i)) #第几行
                        a.append(str(num)) #列数
                        a.append(url)
                        #a = sheet_title,str(i),str(url_colum_num),url
                        #print(a)
                        all_url.append(a)
        else:
            print("[-]"+sheet_title+"此表内URL列存在错误！无法识别")
    if all_url:
        return all_url
        print(all_url)
    else:
        print("无数据或者识别数据出错！")
        exit()

    wb.save(filename=dirname)
    wb.close()


mutex = threading.Lock()
pool = ThreadPoolExecutor(max_workers=50)#线程池设定
def url_reques(duix):#网站url访问测试，检测是否异常
    url = duix[3].strip().replace('https://','').replace('http://','')
    #html_text = ''
    test_url1 = 'http://' + url
    test_url2 = 'https://' + url
    try:
        headers = getheaders()
        res1 = requests.get(test_url1, headers = headers,verify=False,timeout=10)
        code1 = str(res1.status_code)
        flag1 = '1'
    except:
        flag1 = '0'
    if flag1 == '1':
        if code1 in right_code:
            duix.append('正常')
            return duix
        else:
            if int(code1) > 300 & int(code1) < 310:
                rel_url1 =get_real_url(test_url1)
                flag11 = ''
                try:
                    headers = getheaders()
                    res11 = requests.get(rel_url1,headers = headers,verify=False,timeout=10)
                    code11 = str(res11.status_code)
                    flag11 = '1'
                except:
                    flag11 = '0'
                if flag11 == '1':
                    if code11 in right_code:
                        duix.append('正常')
                        return duix
    try:
        headers = getheaders()
        res2 = requests.get(test_url2,headers = headers,verify=False, timeout=10)
        code2 = str(res2.status_code)
        flag2 = '1'
    except:
        duix.append('异常')
        return duix
    if flag2 == '1':
        if code2 in right_code:
            duix.append('正常')
            return duix
        else:
            if int(code2) > 300 & int(code2) < 310:
                rel_url2 = get_real_url(test_url2)
                try:
                    headers = getheaders()
                    res22 = requests.get(rel_url2,headers = headers,verify=False,timeout=10)
                    code22 = str(res22.status_code)
                    flag22 = '1'
                except:
                    flag22 = '0'
                if flag22 == '1':
                    if code22 in right_code:
                        duix.append('正常')
                        return duix
                else:
                    duix.append('异常')
                    return duix

if __name__ == "__main__":
    print("\n*****web应用可用性检测工具 by arno******\n")
    dir_l = input('XLSX文件路径:')
    try:
        all_url = doxlsx(dir_l)
    except:
        print("未识别路径！请退出重新运行！")
        exit()
    print("检测中...")

    data_res_all = []
    tasks = []
    for i in all_url:
        task = pool.submit(url_reques, i)
        #url_reques(i)
        tasks.append(task)
    for a in tasks:
        data_res_all.append(a.result())
    print(data_res_all)
    #wait(pool)
    #time.sleep()
    #print(all_res)
    if data_res_all:
        toxlsx(data_res_all,dir_l)
    else:
        print("写入错误！")
    print("检测已完成！")




